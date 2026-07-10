"""Genera un Excel de prueba (1 año) para importar en app_gestion (FASE 20).

Hojas: productos, gastos, empleados, compras, ventas.
Encabezados == nombres de campo esperados por camposImportacion, para que el
mapeo (manual o por IA) sea directo.
"""

import random
from datetime import date, timedelta

from openpyxl import Workbook

random.seed(42)

ANIO = 2025
INICIO = date(ANIO, 1, 1)
FIN = date(ANIO, 12, 31)


def rango_fechas(inicio, fin):
    dias = (fin - inicio).days + 1
    for i in range(dias):
        yield inicio + timedelta(days=i)


wb = Workbook()

# ---------------------------------------------------------------------------
# productos
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "productos"
ws.append(
    [
        "nombre",
        "categoria",
        "unidad",
        "precioCompra",
        "precioVenta",
        "stockInicial",
        "stockMinimo",
    ]
)

PRODUCTOS = [
    ("Arroz Selecto 5lb", "Granos", "unidad", 180.00, 250.00, 100, 20),
    ("Habichuelas Rojas 1lb", "Granos", "libra", 45.00, 65.00, 80, 15),
    ("Pollo Entero", "Carnes", "libra", 65.00, 95.00, 50, 10),
    ("Salami Embutidos del Cibao", "Embutidos", "unidad", 120.00, 180.00, 40, 10),
    ("Jamon de Pierna", "Embutidos", "libra", 150.00, 220.00, 30, 8),
    ("Longaniza", "Embutidos", "libra", 110.00, 160.00, 25, 5),
    ("Huevos Carton x30", "Lacteos y Huevos", "caja", 220.00, 290.00, 20, 5),
    ("Aceite Vegetal 1gal", "Abarrotes", "galon", 320.00, 420.00, 15, 3),
    ("Leche en Polvo 800g", "Lacteos y Huevos", "unidad", 280.00, 360.00, 25, 5),
    ("Refresco Cola 2L", "Bebidas", "unidad", 90.00, 140.00, 60, 12),
    ("Cerveza Presidente Paca x24", "Bebidas", "paca", 950.00, 1250.00, 20, 4),
    ("Pan de Agua", "Panaderia", "unidad", 8.00, 15.00, 100, 20),
]

for p in PRODUCTOS:
    ws.append(list(p))

# ---------------------------------------------------------------------------
# empleados
# ---------------------------------------------------------------------------
ws = wb.create_sheet("empleados")
ws.append(
    [
        "tipo",
        "nombre",
        "cedula",
        "direccion",
        "telefono",
        "fechaIngreso",
        "salario",
        "frecuenciaPago",
    ]
)

EMPLEADOS = [
    (
        "ventas",
        "Maria Perez",
        "001-1234567-8",
        "Calle Duarte #12, Santiago",
        "809-555-0101",
        "2024-06-01",
        15000.00,
        "quincenal",
    ),
    (
        "ventas",
        "Carlos Gomez",
        "001-2345678-9",
        "Av. Estrella Sadhala #45, Santiago",
        "809-555-0102",
        "2024-09-15",
        14000.00,
        "quincenal",
    ),
    (
        "delivery",
        "Luis Fernandez",
        "001-3456789-0",
        "Calle Sol #8, Santiago",
        "809-555-0103",
        "2025-01-10",
        12000.00,
        "semanal",
    ),
    (
        "ventas",
        "Yesenia Rodriguez",
        "001-4567890-1",
        "Calle Restauracion #20, Santiago",
        "809-555-0104",
        "2023-03-01",
        16000.00,
        "mensual",
    ),
]

for e in EMPLEADOS:
    ws.append(list(e))

# ---------------------------------------------------------------------------
# gastos (recurrentes mensuales + variables)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("gastos")
ws.append(["categoria", "concepto", "fecha", "monto"])

GASTOS_FIJOS = [
    ("Alquiler", "Alquiler del local", 18000.00),
    ("Energia", "Factura de luz (EDE)", 6500.00),
    ("Agua", "Factura de agua (CORAASAN)", 800.00),
    ("Telefono e Internet", "Plan de internet y telefono", 1800.00),
    ("Nomina", "Pago de empacador/ayudante", 8000.00),
]

GASTOS_VARIABLES = [
    ("Transporte", "Flete de mercancia"),
    ("Mantenimiento", "Reparacion de nevera"),
    ("Empaques", "Compra de bolsas y fundas"),
    ("Publicidad", "Volantes y promociones"),
    ("Limpieza", "Articulos de limpieza"),
    ("Impuestos", "Pago de impuestos municipales"),
]

for mes in range(1, 13):
    dia_pago = date(ANIO, mes, 5)
    for categoria, concepto, monto in GASTOS_FIJOS:
        variacion = random.uniform(-0.05, 0.08)
        ws.append(
            [
                categoria,
                concepto,
                dia_pago.isoformat(),
                round(monto * (1 + variacion), 2),
            ]
        )

    # 1 a 3 gastos variables por mes
    for _ in range(random.randint(1, 3)):
        categoria, concepto = random.choice(GASTOS_VARIABLES)
        dia = date(ANIO, mes, random.randint(1, 28))
        monto = round(random.uniform(300, 3500), 2)
        ws.append([categoria, concepto, dia.isoformat(), monto])

# ---------------------------------------------------------------------------
# compras (reabastecimiento semanal)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("compras")
ws.append(
    ["proveedor", "producto", "cantidad", "costoUnitario", "numeroFactura", "fecha"]
)

PROVEEDORES = ["Distribuidora Cibao", "Importadora Nacional", "Suplidora La Economica"]

factura_n = 1
fecha = INICIO
semana = 0
while fecha <= FIN:
    # cada semana, reabastecer 3-4 productos al azar
    productos_semana = random.sample(PRODUCTOS, k=random.randint(3, 4))
    proveedor = PROVEEDORES[semana % len(PROVEEDORES)]
    for (
        nombre,
        _cat,
        _unidad,
        precio_compra,
        _precio_venta,
        _stock,
        _min,
    ) in productos_semana:
        cantidad = random.randint(10, 40)
        costo = round(precio_compra * random.uniform(0.95, 1.05), 2)
        ws.append(
            [
                proveedor,
                nombre,
                cantidad,
                costo,
                f"F-{ANIO}-{factura_n:04d}",
                fecha.isoformat(),
            ]
        )
        factura_n += 1
    fecha += timedelta(days=7)
    semana += 1

# ---------------------------------------------------------------------------
# ventas (2-4 ventas por dia, todo el anio, con repunte en diciembre)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("ventas")
ws.append(["producto", "cantidad", "precioUnitario", "costoUnitario", "fecha"])

for fecha in rango_fechas(INICIO, FIN):
    ventas_dia = random.randint(2, 4)
    if fecha.month == 12:
        ventas_dia += random.randint(1, 3)  # repunte de fin de anio

    for _ in range(ventas_dia):
        nombre, _cat, _unidad, costo_unitario, precio_venta, _stock, _min = (
            random.choice(PRODUCTOS)
        )
        cantidad = (
            round(random.uniform(1, 6), 0)
            if "lb" not in nombre.lower() and "libra" not in _unidad
            else round(random.uniform(0.5, 4), 1)
        )
        precio = round(precio_venta * random.uniform(0.97, 1.03), 2)
        costo = round(costo_unitario * random.uniform(0.97, 1.03), 2)
        ws.append([nombre, cantidad, precio, costo, fecha.isoformat()])

OUT = "test_data/datos_prueba_colmado.xlsx"
wb.save(OUT)
print(f"Generado: {OUT}")
print(f"Filas ventas: {ws.max_row - 1}")
